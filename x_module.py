import datetime
import openpyxl as op
from  openpyxl.styles  import  PatternFill
from openpyxl import workbook
import boto3 #s3 only
import io #s3 only
import boto3

#엑셀 파일 보기 좋게 가공하여 S3 저장하는 로직

class x_main():
    def __init__(self):
        self.s3_bucket='---'
        
    def __main__(self):
        self.xl_load()
        self.actions()
        self.xl_close()
        self.save_s3()
        return self.savetime
    

    def xl_load(self):
        self.xl=op.load_workbook('./rds_result.xlsx') #작업 대상 시트
        self.ws=self.xl.active #활성화 시트
        self.row_max=self.ws.max_row+1
        self.col_max = self.ws.max_column+1

    def actions(self):
        #지정 정보 Patternfill
        for ri in range(3,self.row_max): #row index
            for ci in range(self.col_max-20, self.col_max-12): #col index
                text=self.ws.cell(ri, ci).value
                if text!=None and text!=' ' and text!='':
                    self.ws.cell(ri,ci).fill=PatternFill(fill_type='solid', fgColor="FFFF00")
                    
        #서식 지정
        float_cols=['소수_자료형으로_표시할_컬럼들']
        integer_cols=['정수_자료형으로_표시할_컬럼들']
        accent_cols=['비고']
        for ci in range(1,self.col_max):
            colname=self.ws.cell(2,ci).value
            if colname in float_cols:
                for ri in range(3,self.row_max):
                    self.ws.cell(ri, ci).number_format = '###0.00;[Red](-###0.00)'
            elif colname in integer_cols:
                for ri in range(3,self.row_max):
                    self.ws.cell(ri, ci).number_format='###0;[Red](-###0)'
            elif colname in accent_cols:
                for ri in range(3, self.row_max):
                    text=self.ws.cell(ri, ci).value
                    if text!=None:
                        self.ws.cell(ri,ci).fill=PatternFill(fill_type='solid', fgColor="FFFF00")
                    
    def xl_close(self):
        self.xl.save('./RPA_RESULT.xlsx')
        self.xl.close()
    
    def save_s3(self):
        savetime=str(datetime.datetime.now())
        bucket=self.s3_bucket
        s3=boto3.client('s3')
        s3.upload_file("./RPA_RESULT.xlsx", bucket, f"result_{savetime}.xlsx")
        print(f'RESULT S3 적재 완료')
        self.savetime=savetime
    
        

    
# if __name__ == '__main__':   
    # task=x_main()
    # task.__main__()